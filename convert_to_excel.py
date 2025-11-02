import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
import io

def extract_pokemon_data(json_path):
    """Extrait les données d'un fichier JSON Pokémon"""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # On prend la forme de base (form: 0)
        base_form = None
        for form in data.get('forms', []):
            if form.get('form') == 0:
                base_form = form
                break
        
        if not base_form:
            print(f"⚠️  Pas de forme de base trouvée pour {data.get('dbSymbol', 'inconnu')}")
            return None
        
        # Extraction des données
        pokemon_data = {
            'dbSymbol': data.get('dbSymbol', 'unknown'),
            'sprite_name': base_form.get('resources', {}).get('front', ''),
            'type1': base_form.get('type1', ''),
            'type2': base_form.get('type2', ''),
            'hp': base_form.get('baseHp', 0),
            'atk': base_form.get('baseAtk', 0),
            'def': base_form.get('baseDfe', 0),
            'spa': base_form.get('baseAts', 0),
            'spd_stat': base_form.get('baseDfs', 0),
            'spe': base_form.get('baseSpd', 0),
            'abilities': base_form.get('abilities', [])
        }
        
        return pokemon_data
        
    except json.JSONDecodeError as e:
        print(f"❌ Erreur JSON dans {json_path}: {e}")
        return None
    except Exception as e:
        print(f"❌ Erreur lors de la lecture de {json_path}: {e}")
        return None

def resize_sprite(sprite_path, target_height=60):
    """Redimensionne le sprite pour l'Excel"""
    try:
        img = Image.open(sprite_path)
        # Calculer la largeur proportionnelle
        aspect_ratio = img.width / img.height
        target_width = int(target_height * aspect_ratio)
        img = img.resize((target_width, target_height), Image.Resampling.LANCZOS)
        
        # Sauvegarder en mémoire
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        return img_buffer
    except Exception as e:
        print(f"⚠️  Erreur lors du redimensionnement de {sprite_path}: {e}")
        return None

def load_national_dex(base_path):
    """Charge le fichier national.json et retourne la liste des créatures"""
    national_path = base_path / 'data' / 'national.json'
    
    if not national_path.exists():
        print(f"❌ Le fichier {national_path} n'existe pas!")
        return None
    
    try:
        with open(national_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        creatures = data.get('creatures', [])
        print(f"📖 Pokédex National chargé: {len(creatures)} entrées")
        return creatures
        
    except Exception as e:
        print(f"❌ Erreur lors de la lecture du national.json: {e}")
        return None

def create_pokemon_excel(base_dir='.'):
    """Crée le fichier Excel avec tous les Pokémon"""
    
    base_path = Path(base_dir)
    json_dir = base_path / 'data' / 'pokemon_consolidated'
    sprite_dir = base_path / 'data' / 'pokefront'
    
    # Vérification des dossiers
    if not json_dir.exists():
        print(f"❌ Le dossier {json_dir} n'existe pas!")
        return
    if not sprite_dir.exists():
        print(f"❌ Le dossier {sprite_dir} n'existe pas!")
        return
    
    # Charger le Pokédex National
    creatures = load_national_dex(base_path)
    if creatures is None:
        return
    
    print(f"📂 Dossier JSON: {json_dir}")
    print(f"🖼️  Dossier sprites: {sprite_dir}")
    print()
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pokédex"
    
    # Headers
    headers = ['Sprite', 'Nom', 'Type 1', 'Type 2', 'PV', 'Atq', 'Def', 'Atq Spé', 'Def Spé', 'Vit', 
               'Talent 1', 'Talent 2', 'Talent 3']
    ws.append(headers)
    
    # Style des headers
    from openpyxl.styles import Font
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
    
    # Ajuster la hauteur de la ligne de header
    ws.row_dimensions[1].height = 20
    
    processed = 0
    errors = 0
    row_idx = 2  # On commence à la ligne 2 (après le header)
    
    for creature in creatures:
        db_symbol = creature.get('dbSymbol')
        json_file = json_dir / f"{db_symbol}.json"
        
        if not json_file.exists():
            print(f"❌ Fichier manquant: {json_file.name}")
            errors += 1
            continue
        
        # Extraire les données
        pokemon_data = extract_pokemon_data(json_file)
        if not pokemon_data:
            errors += 1
            continue
        
        # Chercher le sprite
        sprite_name = pokemon_data['sprite_name']
        sprite_path = sprite_dir / f"{sprite_name}.png"
        
        if not sprite_path.exists():
            print(f"⚠️  Sprite manquant pour {db_symbol}: {sprite_name}.png")
            sprite_path = None
        
        # Préparer les données de la ligne
        row_data = [
            '',  # Sprite (sera ajouté après)
            pokemon_data['dbSymbol'],
            pokemon_data['type1'] if pokemon_data['type1'] != '__undef__' else '',
            pokemon_data['type2'] if pokemon_data['type2'] != '__undef__' else '',
            pokemon_data['hp'],
            pokemon_data['atk'],
            pokemon_data['def'],
            pokemon_data['spa'],
            pokemon_data['spd_stat'],
            pokemon_data['spe']
        ]
        
        # Ajouter les talents
        abilities = pokemon_data['abilities']
        for i in range(3):
            if i < len(abilities):
                row_data.append(abilities[i])
            else:
                row_data.append('')
        
        # Écrire la ligne
        ws.append(row_data)
        
        # Ajuster la hauteur de ligne pour le sprite
        ws.row_dimensions[row_idx].height = 45
        
        # Ajouter le sprite si disponible
        if sprite_path:
            try:
                img_buffer = resize_sprite(sprite_path)
                if img_buffer:
                    img = XLImage(img_buffer)
                    # Positionner l'image dans la cellule A
                    cell_ref = f'A{row_idx}'
                    ws.add_image(img, cell_ref)
                else:
                    print(f"⚠️  Impossible de redimensionner le sprite: {db_symbol}")
            except Exception as e:
                print(f"❌ Erreur sprite pour {db_symbol}: {e}")
        
        processed += 1
        row_idx += 1
    
    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 12  # Sprite
    ws.column_dimensions['B'].width = 18  # Nom
    for col in range(3, 9):  # Stats
        ws.column_dimensions[get_column_letter(col)].width = 10
    for col in range(9, 12):  # Talents
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Sauvegarder
    output_file = base_path / 'pokedex_export.xlsx'
    wb.save(output_file)
    
    print("=" * 60)
    print(f"✨ Export terminé!")
    print(f"📁 Fichier créé: {output_file}")
    print(f"✅ Pokémon traités: {processed}")
    if errors > 0:
        print(f"⚠️  Erreurs: {errors}")
    print("=" * 60)

if __name__ == "__main__":
    # Lancer l'extraction depuis le dossier courant
    # Si ton script est à la racine du projet pokedex_web
    create_pokemon_excel('.')
    
    # Si tu lances le script depuis un autre endroit, change le chemin:
    # create_pokemon_excel('/chemin/vers/pokedex_web')